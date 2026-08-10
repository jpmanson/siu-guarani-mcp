from __future__ import annotations

import csv
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .client import GuaraniClient


def normalize_name(value: str) -> str:
    normalized = ''.join(c for c in unicodedata.normalize('NFKD', value or '') if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', normalized).strip().upper()


def merge_alumnos_notas(alumnos: list[dict[str, Any]], notas: list[dict[str, Any]]) -> list[dict[str, Any]]:
    notas_by_name = {normalize_name(row.get('nombre', '')): row for row in notas}
    rows: list[dict[str, Any]] = []
    for idx, alumno in enumerate(sorted(alumnos, key=lambda x: normalize_name(x.get('nombre', ''))), start=1):
        nota = notas_by_name.get(normalize_name(alumno.get('nombre', '')))
        rows.append({
            'Nro': idx,
            'Nombre': alumno.get('nombre', ''),
            'Legajo': alumno.get('legajo', ''),
            'Identificación': nota.get('identificacion', '') if nota else '',
            'Acta': nota.get('acta', '') if nota else '',
            'Fecha promoción': nota.get('fecha_promocion', '') if nota else '',
            'Nota promoción': nota.get('nota_promocion', '') if nota else '',
            'Resultado promoción': nota.get('resultado_promocion_label', '') if nota else '',
            'Observación promoción': nota.get('observacion_promocion', '') if nota else '',
            'Presente asistencia consultada': 'Sí' if alumno.get('presente') else 'No',
            'Hash alumno SIU': alumno.get('alumno_hash', ''),
            'Renglón nota SIU': nota.get('renglon_id', '') if nota else '',
            'Tiene nota cargada': 'Sí' if nota else 'No',
        })
    return rows


def write_exports(rows: list[dict[str, Any]], metadata: dict[str, Any], output_dir: str | Path, basename: str) -> dict[str, Any]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    csv_path = out / f'{basename}.csv'
    xlsx_path = out / f'{basename}.xlsx'
    json_path = out / f'{basename}.json'
    headers = list(rows[0].keys()) if rows else []

    with csv_path.open('w', newline='', encoding='utf-8-sig') as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    json_path.write_text(json.dumps({**metadata, 'rows': rows}, ensure_ascii=False, indent=2), encoding='utf-8')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Estudiantes y notas'
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    widths = [6, 34, 12, 18, 24, 16, 14, 22, 24, 26, 44, 16, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    summary = wb.create_sheet('Resumen')
    for key, value in metadata.items():
        if key != 'rows':
            summary.append([key, value])
    summary.column_dimensions['A'].width = 38
    summary.column_dimensions['B'].width = 90
    for cell in summary['A']:
        cell.font = Font(bold=True)
    wb.save(xlsx_path)

    return {'csv': str(csv_path), 'xlsx': str(xlsx_path), 'json': str(json_path)}


def export_cursada_con_notas(
    zona_clases_url: str,
    zona_comisiones_url: str,
    output_dir: str | Path = 'export',
    basename: str = 'estudiantes_notas_cursada',
) -> dict[str, Any]:
    client = GuaraniClient()
    alumnos = client.alumnos_cursada(zona_clases_url)
    notas = client.notas_cursada(zona_comisiones_url)
    rows = merge_alumnos_notas(alumnos, notas)
    metadata = {
        'generado': datetime.now().isoformat(timespec='seconds'),
        'zona_clases_url': zona_clases_url,
        'zona_comisiones_url': zona_comisiones_url,
        'alumnos_count': len(alumnos),
        'notas_count': len(notas),
        'export_rows_count': len(rows),
        'sin_nota_count': len(rows) - len(notas),
    }
    files = write_exports(rows, metadata, output_dir, basename)
    return {**files, **metadata}
